#coding=utf-8
'''
按姓名排序
'''
import openpyxl
wb1 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\姓名列表.xlsx') # excel文件
sheet = wb1.worksheets[0]
wb3 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\4.xlsx', data_only=True) # excel文件
sheets3 = wb3.worksheets  # 所有sheet表

list1 = []
list2 = []
dic1 = {}
# 将姓名列表中的姓名保存到list1
for i in range(1, sheet.max_row+1):
    if sheet.cell(i,1).value:
        list1.append(sheet.cell(i,1).value)
for i in range(1, sheet.max_row + 1):
    if sheet.cell(i,2).value:
        list2.append(sheet.cell(i,2).value)
list3 = list1 + list2
# 将3.excel内容保存到dic1
for sheet in sheets3:   # 遍历每一个sheet
    for i in range(7, 17):  # 每个sheet的每个村民
        temp_list = []
        for j in range(4, 17):
            temp_list.append(sheet.cell(i, j).value)
        dic1[sheet.cell(i, 3).value] = temp_list

print(len(list1), list1)
print(len(list2), list2)
print(len(list3), list3)
print(len(dic1), dic1)
# 将dic1中数据按顺序填入3.excel
j = 0   # 数据填充的sheet数
index = 7   # 数据填充的行数
for key,value in dic1.items():
    if index == 17:
        index = 7
        j += 1
    if key not in list3:
        sheets3[j].cell(index, 3, key)  # 姓名
        for k in range(4, 17):
            sheets3[j].cell(index, k, value[k-4])
        index += 1
print(j,index)
for key,value in dic1.items():
    if index == 17:
        index = 7
        j += 1
    if key in list1:
        sheets3[j].cell(index, 3, key)  # 姓名
        for k in range(4, 17):
            sheets3[j].cell(index, k, value[k-4])
        index += 1
print(j,index)
for key,value in dic1.items():
    if index == 17:
        index = 7
        j += 1
    if key in list2:
        sheets3[j].cell(index, 3, key)  # 姓名
        for k in range(4, 17):
            sheets3[j].cell(index, k, value[k-4])
        index += 1
print(j,index)
# wb3.save(r'C:\Users\伊海迪\Desktop\4.xlsx') # 修改后要保存

print('杨云海' not in list3)