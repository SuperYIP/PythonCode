#coding=utf-8
'''
python使用openpyxl处理excel数据
    openpyxl库即可以对excel进行读操作也可以进行写操作
'''
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\2.xlsx') # excel文件
sheets = wb.worksheets  # 所有sheet表
value = sheets[0].cell(1,1).value   # 获取某一单元格的值
sheets[0].cell(1,1,'伊海迪')   # 修改某一单元格的值
wb.save(r'C:\Users\伊海迪\Desktop\2.xlsx') # 修改后要保存

