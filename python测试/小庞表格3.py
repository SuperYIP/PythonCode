#coding=utf-8
'''
处理3.excel
'''
import openpyxl
wb3 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\刘德存_处理后.xlsx') # excel文件
sheets3 = wb3.worksheets  # 所有sheet表

for sheet in sheets3:   # 遍历每一个sheet
    for i in range(7, 17):  # 每个sheet的每个村民
        # 计算总作业面积
        a = sheet.cell(i, 6).value if sheet.cell(i, 6).value else 0
        b = sheet.cell(i, 9).value if sheet.cell(i, 9).value else 0
        r1 = a + b
        sheet.cell(i, 12, r1)
        # 计算应收金额合计
        a = sheet.cell(i, 7).value if sheet.cell(i, 7).value else 0
        b = sheet.cell(i, 10).value if sheet.cell(i, 10).value else 0
        r2 = a + b
        sheet.cell(i, 13, r2)
        # 计算实收金额合计
        a = sheet.cell(i, 8).value if sheet.cell(i, 8).value else 0
        b = sheet.cell(i, 11).value if sheet.cell(i, 11).value else 0
        r3 = a + b
        sheet.cell(i, 14, r3)
        # 计算兑付补助金额合计
        r4 = r2 - r3
        sheet.cell(i, 15, r4)

wb3.save(r'C:\Users\伊海迪\Desktop\刘德存_处理后.xlsx') # 修改后要保存
