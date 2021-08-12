#coding=utf-8
import openpyxl
wb1 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\1.xlsx') # excel文件
wb2 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\2.xlsx') # excel文件
wb3 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\3.xlsx') # excel文件

sheets1 = wb1.worksheets  # 所有sheet表
sheets2 = wb2.worksheets  # 所有sheet表
sheets3 = wb3.worksheets  # 所有sheet表
# sheets1_len = len(sheets1)  # 表1的sheet数
# list1 = []
# # 将要合并的数据放入列表list1中
# for sheet in sheets1:   # 遍历每一个sheet
#     for i in range(4, sheet.max_row): # 遍历sheet的每一行
#         list1.append([sheet.cell(i, 2).value, sheet.cell(i, 3).value, sheet.cell(i, 4).value,
#                       sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value])
# for sheet in sheets2:   # 遍历每一个sheet
#     for i in range(4, sheet.max_row): # 遍历sheet的每一行
#         list1.append([sheet.cell(i, 2).value, sheet.cell(i, 3).value, sheet.cell(i, 4).value,
#                       sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value])
dic1 = {}
# 将要合并的数据放入列表list1中
for sheet in sheets1:   # 遍历每一个sheet
    for i in range(4, sheet.max_row): # 遍历sheet的每一行
        dic1[sheet.cell(i, 2).value] = [sheet.cell(i, 3).value, sheet.cell(i, 4).value,
                      sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value]
for sheet in sheets2:   # 遍历每一个sheet
    for i in range(4, sheet.max_row): # 遍历sheet的每一行
        if sheet.cell(i, 2).value in dic1.keys():
            dic1[sheet.cell(i, 2).value] = dic1[sheet.cell(i, 2).value] + \
                                           [sheet.cell(i, 6).value, sheet.cell(i, 7).value, sheet.cell(i, 9).value]
        else:
            dic1[sheet.cell(i, 2).value] = [sheet.cell(i, 3).value, sheet.cell(i, 4).value,
                      sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value]
# 将list1中数据填入3.excel
j = 0   # 数据填充的sheet数
index = 7   # 数据填充的行数
for i in range(len(dic1)):
    if index == 17:
        index = 7
        j += 1
    sheets3[j].cell(index, 3, dic1[i][0])  # 姓名
    sheets3[j].cell(index, 4, dic1[i][1])  # 身份证
    sheets3[j].cell(index, 5, list1[i][2])  # 电话
    if list1[i][3] == '旋耕':
        sheets3[j].cell(index, 6, list1[i][4])  # 旋耕面积
        sheets3[j].cell(index, 7, list1[i][4] * 60) # 应收金额
        sheets3[j].cell(index, 8, list1[i][4] * 42) # 实收金额
    else:
        sheets3[j].cell(index, 9, list1[i][4])
        sheets3[j].cell(index, 10, list1[i][4] * 50)
        sheets3[j].cell(index, 11, list1[i][4] * 35)
    index += 1
# wb3.save(r'C:\Users\伊海迪\Desktop\3.xlsx') # 修改后要保存

# for sheet in sheets3:
#     sheet.cell(7, 3, list1)
# print(list1)
# for temp in sheets1[0].rows:
#     print(temp.value)
# sheets[0].cell(1,1,'伊海迪')   # 修改某一单元格的值
# wb1.save(r'C:\Users\伊海迪\Desktop\2.xlsx') # 修改后要保存
