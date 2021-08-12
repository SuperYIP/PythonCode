#coding=utf-8
'''
1.excel,2.excel-->3.excel
'''
import openpyxl
wb1 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\刘德存.xlsx') # excel文件
wb2 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\刘志涛.xlsx') # excel文件
wb3 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\刘德存_处理后.xlsx') # excel文件
wb4 = openpyxl.load_workbook(r'C:\Users\伊海迪\Desktop\刘志涛_处理后.xlsx') # excel文件

sheets1 = wb1.worksheets  # 所有sheet表
sheets2 = wb2.worksheets  # 所有sheet表
sheets3 = wb3.worksheets  # 所有sheet表
sheets4 = wb4.worksheets  # 所有sheet表
dic1 = {}   # 用字典保存，键为人名，值为
# 将要合并的数据放入列表list1中
for sheet in sheets1:   # 遍历每一个sheet
    for i in range(4, sheet.max_row): # 遍历sheet的每一行
        if sheet.cell(i, 2).value in dic1.keys():
            dic1[sheet.cell(i, 2).value] = dic1[sheet.cell(i, 2).value] + \
                                           [sheet.cell(i, 6).value, sheet.cell(i, 7).value, sheet.cell(i, 9).value]
        else:
            dic1[sheet.cell(i, 2).value] = [sheet.cell(i, 3).value, sheet.cell(i, 4).value,
                      sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value]
for sheet in sheets2:   # 遍历每一个sheet
    for i in range(4, sheet.max_row): # 遍历sheet的每一行
        if sheet.cell(i, 2).value in dic1.keys():
            dic1[sheet.cell(i, 2).value] = dic1[sheet.cell(i, 2).value] + \
                                           [sheet.cell(i, 6).value, sheet.cell(i, 7).value, sheet.cell(i, 9).value]
        else:
            dic1[sheet.cell(i, 2).value] = [sheet.cell(i, 3).value, sheet.cell(i, 4).value,
                      sheet.cell(i, 6).value,sheet.cell(i, 7).value, sheet.cell(i, 9).value]
# 将list1中数据填入3.excel
j = 0   # 数据填充的sheet数
index = 7   # 数据填充的行数
for key,value in dic1.items():
    if index == 17:
        index = 7
        j += 1
    sheets3[j].cell(index, 3, key)  # 姓名
    sheets3[j].cell(index, 4, value[0])  # 身份证
    sheets3[j].cell(index, 5, value[1])  # 电话
    if len(value) == 5:
        if value[2] == '旋耕':
            sheets3[j].cell(index, 6, value[3])  # 旋耕面积
            sheets3[j].cell(index, 7, value[3] * 60) # 应收金额
            sheets3[j].cell(index, 8, value[3] * 42) # 实收金额
        else:
            sheets3[j].cell(index, 9, value[3])
            sheets3[j].cell(index, 10, value[3] * 50)
            sheets3[j].cell(index, 11, value[3] * 35)
    else:
        sheets3[j].cell(index, 6, value[3])  # 旋耕面积
        sheets3[j].cell(index, 7, value[3] * 60) # 应收金额
        sheets3[j].cell(index, 8, value[3] * 42) # 实收金额
        sheets3[j].cell(index, 9, value[3])
        sheets3[j].cell(index, 10, value[3] * 50)
        sheets3[j].cell(index, 11, value[3] * 35)
    index += 1
wb3.save(r'C:\Users\伊海迪\Desktop\3.xlsx') # 修改后要保存

